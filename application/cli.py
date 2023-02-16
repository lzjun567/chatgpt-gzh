"""
命令行工具
"""
import logging
import random
from typing import List

import click
from faker import Faker
from flask import current_app
from openpyxl import load_workbook

from application import StudentUser
from application.enum_field import MaterialType
from application.extensions import db
from application.models import InstitutionUser, Institution, User, RankInvolve, MaterialCategory
from application.models.permission import PermissionTemplate
from application.services import rank_service, micropage_service
from application.tasks import stats_task

logger = logging.getLogger()
logger.setLevel(level=logging.INFO)
fake = Faker("zh_CN")


def configure_cli(app):
    @app.cli.command("stat")
    def stst():
        stats_task.stat()

    @app.cli.command("add_institution")
    @click.option("--name", prompt="机构名称")
    @click.option("--creator_name", prompt="创建者姓名")
    @click.option("--creator_phone", prompt="创建者手机号")
    def add_institution(name, creator_name, creator_phone):
        Institution.create(name=name, creator_name=creator_name, creator_phone=creator_phone)

    @app.cli.command("sync_institution_user_to_platform")
    def sync_institution_user_to_platform():
        """
        同步机构用户到营销平台
        """
        from application.extensions import authapi
        for item in InstitutionUser.query.all():
            platform_userid = authapi.add_user(phone=item.user.phone, password=item.user.password, name=item.user.name)
            item.user.platform_userid = platform_userid
            print(item.user.name, platform_userid)
            db.session.commit()

    @app.cli.command("import_permission")
    def import_permission():
        # 下标从1开始
        wb = load_workbook("权限code表.xlsx")
        sheet = wb.active  # 当前的sheet
        sheet = wb.get_sheet_by_name("患教管理后台code")
        current_app.logger.info("导入：%s列*%s行", sheet.max_column, sheet.max_row)
        start_row = 2
        # 添加root权限
        root_permission = PermissionTemplate.get_by_code("root")
        if not root_permission:
            root_permission = PermissionTemplate(code="root", name="root")
            db.session.add(root_permission)
            db.session.commit()

        level_1_name = None
        level_1_code = None

        level_2_name = None
        level_2_code = None

        for i in range(start_row, sheet.max_row + 1):
            level_1_name = (sheet.cell(i, 1).value or "").strip() or level_1_name
            level_1_code = (sheet.cell(i, 2).value or "").strip() or level_1_code
            if level_1_code:
                permission_level_1 = PermissionTemplate.query.filter(PermissionTemplate.code == level_1_code).first()
                if not permission_level_1:
                    permission_level_1 = PermissionTemplate(name=level_1_name,
                                                            code=level_1_code,
                                                            level=1,
                                                            parent_id=root_permission.id,
                                                            root_id=root_permission.id,
                                                            key=root_permission.key + root_permission.id + "-")
                    db.session.add(permission_level_1)
                    print("新增模块：", level_1_name)
                else:
                    permission_level_1.name = level_1_name
                db.session.commit()

                level_2_name = (sheet.cell(i, 3).value or "").strip() or level_2_name
                level_2_code = (sheet.cell(i, 4).value or "").strip() or level_2_code

                if level_2_code:
                    permission_level_2 = PermissionTemplate.get_by_code(level_2_code)
                    if not permission_level_2:
                        permission_level_2 = PermissionTemplate(name=level_2_name,
                                                                code=level_2_code,
                                                                level=2,
                                                                parent_id=permission_level_1.id,
                                                                root_id=root_permission.id,
                                                                key=permission_level_1.key + permission_level_1.id + "-")
                        db.session.add(permission_level_2)
                        print("新增模块：", level_2_name)
                    else:
                        permission_level_2.name = level_2_name
                    db.session.commit()
                    level_3_name = (sheet.cell(i, 5).value or "").strip()
                    level_3_code = (sheet.cell(i, 6).value or "").strip()
                    if level_3_code:
                        print(level_3_code)
                        permission_level_3 = PermissionTemplate.get_by_code(level_3_code)
                        if not permission_level_3:
                            permission_level_3 = PermissionTemplate(name=level_3_name,
                                                                    code=level_3_code,
                                                                    level=3,
                                                                    parent_id=permission_level_2.id,
                                                                    root_id=root_permission.id,
                                                                    key=permission_level_2.key + permission_level_2.id + "-")
                            db.session.add(permission_level_3)
                            print("新增模块：", level_3_name)
                        else:
                            permission_level_3.name = level_3_name
                        db.session.commit()

    @app.cli.command("fix_user_id")
    def fix_user_id():
        users: List[StudentUser] = StudentUser.query.filter(StudentUser.user_id == '62cd31c0c890dc77eb7055c1',
                                                            StudentUser.name != "志军").all()
        for student_user in users:
            # print(student_user.id, student_user.name)
            user = User(phone=None, name=student_user.name, avatar=student_user.avatar)
            db.session.add(user)
            db.session.commit()
            student_user.user_id = user.id
            db.session.commit()

    @app.cli.command("insert_test_rank")
    def insert_test_rank():
        rank_id = "6386ca05f43a240008edb453"
        institution_id = "61d8036f6dfb5130645837cb"
        users: List[StudentUser] = StudentUser.query.filter(
            StudentUser.institution_id == institution_id).limit(60).all()
        for user in users:
            obj = RankInvolve.upsert(filter_data=dict(user_id=user.id,
                                                      rank_id=rank_id),
                                     update_data={})
            db.session.add(obj)
            db.session.commit()

        involves: List[RankInvolve] = RankInvolve.query.filter(RankInvolve.rank_id == rank_id).all()
        received_users: List[StudentUser] = StudentUser.query.filter(
            StudentUser.institution_id == institution_id).limit(500).all()
        for u in received_users:
            rank_service.receive_invite(random.choice(involves).id, u)

    @app.cli.command("set_audio_material")
    def default_audio_material():
        # 指定音频默认分类
        for institution in Institution.query.filter(Institution.is_deleted.is_(False)).all():
            # 音频分类
            audio_mc = MaterialCategory(material_num=0,
                                        type=MaterialType.AUDIO.value,
                                        is_default=True,
                                        name="默认分组",
                                        institution_id=institution.id,
                                        )
            db.session.add(audio_mc)
            db.session.commit()

    @app.cli.command("default_recycle_material")
    def default_recycle_material():
        # 指定回收站默认分类
        for institution in Institution.query.filter(Institution.is_deleted.is_(False)).all():
            # 音频分类
            audio_mc = MaterialCategory.upsert(filter_data=dict(name="回收站",
                                                                institution_id=institution.id,
                                                                type=MaterialType.VIDEO.value, ),
                                               update_data=dict(material_num=0,
                                                                type=MaterialType.VIDEO.value,
                                                                is_default=False,
                                                                name="回收站",
                                                                is_recycle=1,
                                                                institution_id=institution.id, )
                                               )
            db.session.add(audio_mc)
            db.session.commit()
